import re
import os
import openpyxl
import scipy.io as sio
import matplotlib.pyplot as plt
import pandas as pd

plt.style.use('bmh')
plt.rcParams["font.family"] = "Times New Roman"
# plt.rcParams["font.size"] = 20
plt.rcParams['axes.unicode_minus'] = True
plt.rcParams['text.color'] = 'black'
plt.set_cmap('jet')
colors = ['blue', 'orange', 'red', 'forestgreen', 'darkviolet', ]


def _crypt(index):
    def read_xlsx_to_dict(file_path):
        workbook = openpyxl.load_workbook(file_path)
        datas = {}

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # 判断当前工作表是否为空
            if sheet.max_row <= 1:
                continue

            if index == 1:
                sheet_name_norm = sheet_name.split('.')[0]
            else:
                sheet_name_norm = sheet_name[:-1]

            # 只读取 duration
            datas[sheet_name_norm] = []

            # 获取表头
            headers = [cell.value for cell in sheet[1]]

            # 遍历每一行数据
            for row in range(2, sheet.max_row + 1):
                row_data = [cell.value for cell in sheet[row]]

                # 将数据存入字典
                for col, value in enumerate(row_data):
                    column_name = headers[col]
                    if column_name == 'duration':
                        datas[sheet_name_norm].append(value)

        return datas

    # 使用示例
    datas = read_xlsx_to_dict(f'./raw_data/Crypt_{index}.xlsx')
    return datas


def _plt_crypt(datas, n_node, root_path):
    save_path = root_path + f'crypt/'
    if os.path.exists(save_path) is False:
        os.mkdir(save_path)
    keys = datas['1'].keys()
    for key in keys:
        phase, mode, key_len, size = key.split('_')
        phase = 'Encryption' if phase == 'EN' else 'Decryption'
        key_len = str(int(key_len) * 8)
        size = str(int(size) * 16) + ' Bytes'
        batch = f'{phase}-{mode}-{key_len}-{size}'
        # # sole
        # for i in range(n_node):
        #     plt.figure()
        #     plt.plot(datas[str(i + 1)][key], color=colors[i])
        #     plt.title(batch)
        #     plt.xlabel('Time slot')
        #     plt.ylabel('Latency (ms)')
        #     plt.tight_layout()
        #     plt.savefig(save_path + f'{batch}.png')
        #     plt.close()
        #     # plt.show()
        # all
        plt.figure()
        for i in range(n_node):
            plt.plot(datas[str(i + 1)][key], label=f'Node {i + 1}', color=colors[i])
        plt.title(batch)
        plt.xlabel('Time slot')
        plt.ylabel('Latency (ms)')
        plt.legend()
        plt.tight_layout()

        plt.savefig(root_path + f'{batch}.png')
        plt.savefig(save_path + f'{batch}.png')
        plt.show()
        plt.close()


def _iw(index, timestamp):
    def process_value(value):
        # 分割字符串以获取数值和单位
        pattern = r"(\s*[\d\.\-]+\s*)\s*\[.*?\]\s*(\w+)"
        # 搜索匹配项
        match = re.search(pattern, value)
        if match:
            number = match.group(1).strip()
            # 第二个捕获组是单位
            unit = '(' + match.group(2) + ')'
        else:
            number = value
            unit = ''
        return number, unit

    # 读取并解析文本文件
    def parse_data_from_file(file_path):
        data_dict = {}
        with open(file_path, 'r') as file:
            for line in file:
                # 提取
                key, sep, value = line.partition(':')
                if sep == '':  # 如果 sep 是空的，说明没有找到冒号，跳过这行
                    continue
                key = key.strip()
                value = value.strip()

                special_keys = ['signal', 'signal avg']
                if key in special_keys:
                    value, unit = process_value(value)
                    key = f"{key} {unit}"  # 创建新的键
                else:
                    # 处理，使用正则表达式匹配数值和单位
                    match = re.match(r'(\d+\.?\d*)\s*([a-zA-Z]+)', value)
                    if match:
                        # 如果匹配成功，提取数值和单位
                        number = float(match.group(1))  # 转换数值为浮点数
                        unit = match.group(2)  # 提取单位
                        # 修改键，添加单位
                        key += f" ({unit})"
                        # 更新值，只保留数值
                        value = number
                    else:
                        # 如果没有匹配到单位，尝试直接转换为整数或浮点数
                        try:
                            value = int(value)
                        except ValueError:
                            try:
                                value = float(value)
                            except ValueError:
                                pass  # 如果转换失败，保持为字符串

                # 存入列表
                if key in data_dict:
                    # 如果键已经存在，将值添加到列表中
                    data_dict[key].append(value)
                else:
                    # 如果键不存在，直接赋值
                    data_dict[key] = [value]
        return data_dict

    # 读取数据文件
    iw_file_path = f"iw_output_{timestamp}_{index}"
    data_dict = parse_data_from_file(iw_file_path + ".txt")

    # 创建一个空的DataFrame
    df = pd.DataFrame()

    # 遍历字典，将每个键和对应的列表添加到DataFrame中
    for idx, (key, values) in enumerate(data_dict.items()):
        # 创建一个新的Series，其中索引是列表值，索引名是键
        series = pd.Series(values, name=key)
        # 将Series添加到DataFrame中
        df[key] = series

    # 文件名
    xlsx_file_path = f"iw_output_{timestamp}" + ".xlsx"
    mode = 'a' if os.path.exists(xlsx_file_path) else 'w'

    # 使用ExcelWriter上下文管理器
    with pd.ExcelWriter(xlsx_file_path, mode=mode, engine='openpyxl', if_sheet_exists='replace') as writer:
        # 将DataFrame写入对应的sheet中
        df.to_excel(writer, sheet_name=f"Node {index}", index=False)

    # # 保存为xlsx文件
    # df.to_excel(iw_file_path + ".xlsx", index=False)
    #
    # # 保存为mat文件
    # savemat("data.mat", {"data": df})

    print(f"Node {index} iw 数据已成功保存。")

    return data_dict


def _iw_xlsx(index):
    df = pd.read_excel(f"./raw_data/iw_output_2024-05-24.xlsx", sheet_name=f"Node {index}")
    data_dict = {}
    for column in df.columns:
        data_dict[column] = df[column].tolist()
    return data_dict


def _ping(index, timestamp):
    # 读取txt文件内容
    def read_ping_output(file_path):
        with open(file_path, 'r') as file:
            ping_output = file.readlines()
        return ping_output

    # 解析函数
    def parse_ping_output(lines):
        # 正则表达式匹配 rtt 和 time
        rtt_pattern = re.compile(r'([\d\.]+)/([\d\.]+)/([\d\.]+)/([\d\.]+)')
        time_pattern = re.compile(r'time=([\d\.]+) ms')

        ping_data_list = {'rtt_min': [], 'rtt_avg': [], 'rtt_max': [], 'rtt_mdev': [], 'latency': [], 'packet_loss': []}

        for line in lines:
            # 匹配 rtt
            rtt_match = rtt_pattern.search(line)
            if rtt_match:
                rtt_min, rtt_avg, rtt_max, rtt_mdev = map(float, rtt_match.groups())
                ping_data_list['rtt_min'].append(rtt_min)
                ping_data_list['rtt_avg'].append(rtt_avg)
                ping_data_list['rtt_max'].append(rtt_max)
                ping_data_list['rtt_mdev'].append(rtt_mdev)

            # 匹配 time
            time_match = time_pattern.search(line)
            if time_match:
                time_value = float(time_match.group(1))
                ping_data_list['latency'].append(time_value)

            # 匹配 packet loss
            loss_pattern = re.compile(r'([0-9]+)% packet loss')
            loss_match = loss_pattern.search(line)
            if loss_match:
                packet_loss = int(loss_match.group(1))
                ping_data_list['packet_loss'] = packet_loss

        return ping_data_list

    # 读取文件并解析数据
    file_path = f'./raw_data/ping_output_2024-05-24_{index}.txt'  # 替换为你的txt文件路径
    if not os.path.exists(file_path):
        print(f"文件 {file_path} 不存在。")
        return
    ping_data_list = parse_ping_output(read_ping_output(file_path))

    # 将解析后的数据转换为DataFrame
    df = pd.DataFrame(ping_data_list)

    # 写入xlsx文件
    output_excel_path = 'ping_statistics.xlsx'  # 你想要保存的xlsx文件路径
    df.to_excel(output_excel_path, index=False)

    print(f"Node {index} 数据已成功写入到 {output_excel_path} 文件中。")

    return ping_data_list


def _plot(cmd, datas, keys, n_node, root_path):
    save_path = root_path + f'{cmd}/'
    if not os.path.exists(save_path):
        os.makedirs(save_path)
    for key in keys:
        try:
            if isinstance(int(datas['1'][key][0]), int):
                # 数据是整数类型
                print('{} is an int'.format(key))
            elif isinstance(float(datas['1'][key][0]), float):
                # 数据是浮点数类型
                print('{} is a float'.format(key))
        except (ValueError, TypeError):
            # 数据不是整数或浮点数类型
            print('{} is a {}'.format(key, type(datas['1'][key][0])))
            continue
        # sole
        for i in range(n_node):
            plt.figure()
            plt.plot(datas[str(i + 1)][key], color=colors[i])
            plt.title(f'{cmd} - Node {i + 1}')
            plt.xlabel('Time slot')
            y_label = key if key != 'latency' else 'Latency (ms)'
            plt.ylabel(y_label)
            plt.tight_layout()
            plt.savefig(save_path + f'{cmd}_{key}_{i + 1}.png')
            plt.close()
            # plt.show()
        # all
        plt.figure()
        for i in range(n_node):
            plt.plot(datas[str(i + 1)][key], label=f'Node {i + 1}', color=colors[i])
        plt.title(cmd)
        plt.xlabel('Time slot')
        y_label = key if key != 'latency' else 'Latency (ms)'
        plt.ylabel(y_label)
        plt.legend()
        plt.tight_layout()
        if cmd == 'ping':
            plt.savefig(root_path + f'{cmd}_{key}.png')
            plt.savefig(save_path + f'{cmd}_{key}.png')
        else:
            plt.savefig(root_path + f'{key}.png')
            plt.savefig(save_path + f'{key}.png')
        plt.show()
        plt.close()


if __name__ == '__main__':
    from datetime import datetime

    # timestamp = datetime.now().strftime('%Y-%m-%d')
    node_num = 5
    timestamp_collect = '2024-05-24'
    file_path = './summary/'

    ping_data = {}
    iw_data = {}
    crypt_data = {}
    for i in range(1, 1 + node_num):
        # ping_data[f"{i}"] = _ping(i, timestamp_collect)
        # iw_data[f"{i}"] = _iw(i, timestamp_collect)
        crypt_data[f"{i}"] = _crypt(i)

    _plt_crypt(crypt_data, 5, file_path)

    # ping_keys = ['latency']
    # _plot('ping', ping_data, ping_keys, 5, file_path)
    #
    # iw_keys = [
    #     "Station 0a",
    #     "inactive time (ms)",
    #     "rx bytes",
    #     "rx packets",
    #     "tx bytes",
    #     "tx packets",
    #     "tx retries",
    #     "tx failed",
    #     "beacon loss",
    #     "beacon rx",
    #     "rx drop misc",
    #     "signal (dBm)",
    #     "signal avg (dBm)",
    #     "beacon signal avg (dBm)",
    #     "tx bitrate (MBit)",
    #     "rx bitrate (MBit)",
    #     "authorized",
    #     "authenticated",
    #     "associated",
    #     "preamble",
    #     "WMM/WME",
    #     "MFP",
    #     "TDLS peer",
    #     "DTIM period",
    #     "beacon interval",
    #     "short preamble",
    #     "short slot time",
    #     "connected time (seconds)",
    #     "addr 30",
    #     "channel 153 (5765 MHz), width (MHz)",
    #     "Connected to 0a",
    #     "SSID",
    #     "freq",
    #     "RX (bytes)",
    #     "TX (bytes)",
    #     "signal",
    #     "bss flags",
    #     "dtim period",
    #     "beacon int"
    # ]
    # _plot('iw', iw_data, iw_keys, 5, file_path)
